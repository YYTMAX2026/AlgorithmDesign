"""
生成实验数据Excel文件
包含：排序比较次数统计、背包问题结果、1000个物品数据
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

wb = Workbook()

# ===================================================
# Sheet 1: 排序比较次数
# ===================================================
ws1 = wb.active
ws1.title = "排序算法比较次数"

headers = ['输入规模 N', '冒泡排序比较次数', '归并排序比较次数', '快速排序比较次数', '归并子问题数', '快速排序子问题数']
ws1.append(headers)

with open('D:/AlgorithmDesign/data/sorting_results.csv', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        ws1.append([int(row['N']), int(row['BubbleCmp']), int(row['MergeCmp']),
                    int(row['QuickCmp']), int(row['MergeSubproblems']), int(row['QuickSubproblems'])])

# 表头样式
header_fill = PatternFill(fill_type='solid', fgColor='4472C4')
for cell in ws1[1]:
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 列宽
ws1.column_dimensions['A'].width = 14
for col in ['B','C','D','E','F']:
    ws1.column_dimensions[col].width = 18

# 添加折线图
chart1 = LineChart()
chart1.title = "排序算法比较次数对比"
chart1.y_axis.title = "比较次数"
chart1.x_axis.title = "输入规模 N"
chart1.width = 20; chart1.height = 14

cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
for col_idx, name in [(2,'冒泡排序'), (3,'归并排序'), (4,'快速排序')]:
    data = Reference(ws1, min_col=col_idx, min_row=1, max_row=ws1.max_row)
    chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws1.add_chart(chart1, "H2")

# ===================================================
# Sheet 2: 背包问题执行时间
# ===================================================
ws2 = wb.create_sheet("背包问题执行时间")
ws2.append(['算法', '物品数量N', '背包容量C', '最优总价值', '总重量', '选中物品数', '执行时间(ms)'])

with open('D:/AlgorithmDesign/data/knapsack_results.csv', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        algo_cn = {'DynamicProgramming':'动态规划', 'Greedy':'贪心法', 'Backtrack':'回溯法'}.get(row['Algorithm'], row['Algorithm'])
        ws2.append([algo_cn, int(row['N']), int(row['Capacity']),
                    float(row['TotalValue']), int(row['TotalWeight']),
                    int(row['SelectedCount']), int(row['TimeMs'])])

header_fill2 = PatternFill(fill_type='solid', fgColor='70AD47')
for cell in ws2[1]:
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = header_fill2
    cell.alignment = Alignment(horizontal='center', vertical='center')

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 14

# ===================================================
# Sheet 3: 1000个物品详细数据（表1）
# ===================================================
ws3 = wb.create_sheet("1000物品数据")
ws3.append(['物品编号', '物品重量', '物品价值'])

with open('D:/AlgorithmDesign/data/items_1000.csv', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        ws3.append([int(row['物品编号']), int(row['物品重量']), float(row['物品价值'])])

header_fill3 = PatternFill(fill_type='solid', fgColor='ED7D31')
for cell in ws3[1]:
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = header_fill3
    cell.alignment = Alignment(horizontal='center')

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 14

wb.save('D:/AlgorithmDesign/data/实验数据.xlsx')
print('Excel文件已保存：D:/AlgorithmDesign/data/实验数据.xlsx')
